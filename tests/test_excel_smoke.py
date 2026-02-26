"""Smoke test: minimal xlsx with 'Задание №1'..'Задание №13', parser finds blocks."""
import tempfile
from pathlib import Path
import pytest
from openpyxl import Workbook

from app.core.excel.blocks import find_task_blocks
from app.core.excel.table_detect import detect_tables_in_block
from app.core.excel.importer import parse_workbook


def test_blocks_found():
    wb = Workbook()
    ws = wb.active
    if not ws:
        pytest.skip("no active sheet")
    for i in range(1, 14):
        ws.cell(row=i * 2, column=1, value=f"Задание №{i}")
        ws.cell(row=i * 2 + 1, column=1, value="some content")
    blocks = find_task_blocks(ws)
    assert len(blocks) == 13
    nums = [b.task_num for b in blocks]
    assert nums == list(range(1, 14))


def test_parse_workbook_smoke():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb = Workbook()
        ws = wb.active
        if not ws:
            pytest.skip("no active sheet")
        # Строка с "Задание №N" пропускается (якорь); таблица — со следующей строки
        for i in range(1, 14):
            anchor_row = i * 2
            ws.cell(row=anchor_row, column=1, value=f"Задание №{i}")
            ws.cell(row=anchor_row + 1, column=1, value="A")
            ws.cell(row=anchor_row + 1, column=2, value="B")
        wb.save(path)
        wb.close()
        parsed = parse_workbook(path)
        assert len(parsed.tasks) == 13
        assert 1 in parsed.tasks
        t1 = parsed.tasks[1]
        assert t1.tables or t1.text_lines
    finally:
        Path(path).unlink(missing_ok=True)


def test_parse_workbook_with_separator_row():
    """Parser and task 3 logic: separator row (dots only) is skipped."""
    from app.core.checks.common import is_separator_row

    assert is_separator_row([".....", "…", ""], max_col=3) is True
    assert is_separator_row(["1", "2"], max_col=2) is False
