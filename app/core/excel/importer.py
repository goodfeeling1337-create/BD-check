"""Parse workbook into ParsedSolution (task blocks -> tables/text)."""
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import load_workbook

from app.core.excel.blocks import TaskBlock, find_task_blocks
from app.core.excel.table_detect import TableInBlock, detect_tables_in_block


@dataclass
class ExtractedTable:
    """One table: header list (labels), rows of cell values."""
    headers: list[str]
    rows: list[list[Any]]


@dataclass
class TaskContent:
    """Content of one task block: tables and/or raw text lines."""
    task_num: int
    tables: list[ExtractedTable] = field(default_factory=list)
    text_lines: list[str] = field(default_factory=list)
    raw_cell_grid: list[list[Any]] = field(default_factory=list)  # optional full block dump


@dataclass
class ParsedSolution:
    """Full parsed solution: task 1..13 content."""
    tasks: dict[int, TaskContent] = field(default_factory=dict)
    sheet_name: str = ""


def _cell_value(ws, row: int, col: int) -> Any:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return v


def _normalize_cell_display(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if isinstance(val, (int, float)) and float(val) == int(val):
        return str(int(val))
    return s


def _extract_table(ws, table: TableInBlock) -> ExtractedTable:
    headers = []
    for c in range(1, table.max_col + 1):
        v = _cell_value(ws, table.header_row, c)
        headers.append(str(v).strip() if v is not None else "")
    rows = []
    for r in table.data_rows:
        row = []
        for c in range(1, table.max_col + 1):
            v = _cell_value(ws, r, c)
            row.append(_normalize_cell_display(v))
        rows.append(row)
    return ExtractedTable(headers=headers, rows=rows)


def _line_looks_like_instruction(line: str) -> bool:
    """Отсекаем строки-инструкции из text_lines."""
    import re
    s = line.strip()
    if not s:
        return True
    if re.search(r"Задание\s*№\s*\d+", s, re.IGNORECASE):
        return True
    if re.match(r"^\s*ответ\s*:?\s*$", s, re.IGNORECASE):
        return True
    if re.match(r"^\s*ответ\s*:?\s*[^a-zA-Zа-яА-Я0-9_]{0,5}", s, re.IGNORECASE) and len(s) < 80:
        return True
    return False


def _block_text_lines(
    ws, start_row: int, end_row: int, max_col: int = 30, skip_rows: Optional[set[int]] = None
) -> list[str]:
    """Собирает текст по строкам блока. skip_rows — номера строк (1-based) не включать (якорь, инструкции)."""
    skip_rows = skip_rows or set()
    lines = []
    for r in range(start_row, end_row):
        if r in skip_rows:
            continue
        parts = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        if parts:
            raw = " ".join(parts)
            if not _line_looks_like_instruction(raw):
                lines.append(raw)
    return lines


def _parse_block(ws, block: TaskBlock) -> TaskContent:
    tables = detect_tables_in_block(
        ws, block.start_row, block.end_row, anchor_row=block.anchor_row
    )
    extracted_tables = [_extract_table(ws, t) for t in tables]
    # Не включать строку-якорь в text_lines, чтобы не попадали "ответ:" и подсказки
    text_lines = _block_text_lines(
        ws, block.start_row, block.end_row, skip_rows={block.anchor_row}
    )
    return TaskContent(
        task_num=block.task_num,
        tables=extracted_tables,
        text_lines=text_lines,
    )


def parse_workbook(path: Union[str, Path]) -> ParsedSolution:
    """Load Excel file and parse first sheet into task blocks and content."""
    path = Path(path)
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return ParsedSolution(sheet_name="")
    name = ws.title
    blocks = find_task_blocks(ws)
    tasks = {}
    for b in blocks:
        tasks[b.task_num] = _parse_block(ws, b)
    wb.close()
    return ParsedSolution(tasks=tasks, sheet_name=name)
