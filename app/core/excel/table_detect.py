"""Detect tables (header + data rows) within a task block."""
import re
from dataclasses import dataclass
from typing import TYPE_CHECKING, Optional

from app.core.checks.common import SEPARATOR_ROW_RE

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# Строка с "Задание №N" или "ответ:" — инструкция, не данные
TASK_ANCHOR_RE = re.compile(r"Задание\s*№\s*\d+", re.IGNORECASE)
ANSWER_PREFIX_RE = re.compile(r"^\s*ответ\s*:?\s*", re.IGNORECASE)


@dataclass
class TableInBlock:
    """Таблица: строка заголовка, строки данных, диапазон колонок (таблица может быть сдвинута)."""
    header_row: int
    data_rows: list[int]
    min_col: int
    max_col: int


def _row_is_separator(ws: "Worksheet", row: int, min_col: int, max_col: int) -> bool:
    """True if row looks like separator (only dots/dashes/ellipsis/empty). Uses SEPARATOR_ROW_RE."""
    for c in range(min_col, max_col + 1):
        val = ws.cell(row=row, column=c).value
        s = (str(val).strip() if val is not None else "")
        if not s:
            continue
        if not SEPARATOR_ROW_RE.match(s):
            return False
    return True


def _is_instruction_row(ws: "Worksheet", row: int, max_col: int = 30) -> bool:
    """True if row looks like instruction (task title, "ответ:", long hint), not data."""
    cells = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip():
            cells.append(str(v).strip())
    if not cells:
        return True
    full_text = " ".join(cells)
    # Строка с "Задание №"
    if TASK_ANCHOR_RE.search(full_text):
        return True
    # Первая ячейка — только "ответ:" или начинается с "ответ:"
    if cells and ANSWER_PREFIX_RE.match(cells[0]) and len(cells) <= 2:
        return True
    if cells and cells[0].lower().strip().startswith("ответ") and len(cells) <= 3:
        return True
    # Одна длинная подсказка (например "порядок столбцов лучше изменить...")
    if len(cells) <= 2 and any(len(s) > 50 for s in cells):
        return True
    return False


def _is_header_like(ws: "Worksheet", row: int, max_col: int) -> bool:
    """Heuristic: row has several non-empty string cells and is not instruction."""
    non_empty = 0
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip():
            non_empty += 1
    return non_empty >= 2


def detect_tables_in_block(
    ws: "Worksheet",
    start_row: int,
    end_row: int,
    max_col: int = 30,
    anchor_row: Optional[int] = None,
) -> list[TableInBlock]:
    """
    Within [start_row, end_row) find tables: each has one header row and consecutive data rows.
    Skip separator rows, anchor row (Задание №N), and instruction rows (ответ:, long hints).
    """
    result: list[TableInBlock] = []
    r = start_row
    while r < end_row:
        # Пропуск строки-якоря (Задание №N)
        if anchor_row is not None and r == anchor_row:
            r += 1
            continue
        # Skip separator rows
        if _row_is_separator(ws, r, 1, max_col):
            r += 1
            continue
        # Строки-инструкции не считать шапкой таблицы
        if _is_instruction_row(ws, r, max_col):
            r += 1
            continue
        if not _is_header_like(ws, r, max_col):
            r += 1
            continue
        header_row = r
        min_col = max_col + 1
        cols_used = 0
        for c in range(1, max_col + 1):
            v = ws.cell(row=header_row, column=c).value
            if v is not None and str(v).strip():
                if min_col > c:
                    min_col = c
                cols_used = c
        min_col = min(min_col, cols_used) if cols_used else 1
        cols_used = max(cols_used, 1)
        data_rows_list: list[int] = []
        r += 1
        while r < end_row:
            if _row_is_separator(ws, r, min_col, cols_used):
                r += 1
                continue
            if _is_instruction_row(ws, r, max_col):
                r += 1
                continue
            has_content = any(
                ws.cell(row=r, column=c).value is not None
                and str(ws.cell(row=r, column=c).value).strip()
                for c in range(min_col, cols_used + 1)
            )
            if has_content:
                data_rows_list.append(r)
                r += 1
            else:
                break
        result.append(
            TableInBlock(
                header_row=header_row,
                data_rows=data_rows_list,
                min_col=min_col,
                max_col=cols_used,
            )
        )
        r += 1
    return result
