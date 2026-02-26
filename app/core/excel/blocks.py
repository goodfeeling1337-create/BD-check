"""Find task blocks by anchor 'Задание №<number>'."""
import re
from dataclasses import dataclass
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

TASK_ANCHOR_RE = re.compile(r"Задание\s*№\s*(\d+)", re.IGNORECASE)


@dataclass
class TaskBlock:
    """Block of cells for one task (1..13)."""
    task_num: int
    start_row: int  # 1-based, inclusive
    end_row: int    # 1-based, exclusive (first row of next task or end of sheet)
    anchor_row: int  # row where "Задание №N" was found


def find_task_blocks(ws: "Worksheet") -> list[TaskBlock]:
    """
    Scan worksheet for 'Задание №N' and build blocks.
    Block for task N: from anchor row (inclusive) to next anchor or end of sheet.
    """
    anchors: list[tuple[int, int]] = []  # (row, task_num)
    for row_idx in range(1, ws.max_row + 1):
        row_text = " ".join(
            str(ws.cell(row=row_idx, column=c).value or "").strip()
            for c in range(1, min(ws.max_column + 1, 50))
        )
        m = TASK_ANCHOR_RE.search(row_text)
        if m:
            num = int(m.group(1))
            if 1 <= num <= 13:
                anchors.append((row_idx, num))

    # Sort by row; dedupe by task_num (keep first occurrence per task)
    seen: set[int] = set()
    ordered: list[tuple[int, int]] = []
    for row, num in sorted(anchors, key=lambda x: (x[0], x[1])):
        if num not in seen:
            seen.add(num)
            ordered.append((row, num))

    blocks: list[TaskBlock] = []
    for i, (row, num) in enumerate(ordered):
        end = ordered[i + 1][0] if i + 1 < len(ordered) else ws.max_row + 1
        blocks.append(TaskBlock(task_num=num, start_row=row, end_row=end, anchor_row=row))
    return blocks
